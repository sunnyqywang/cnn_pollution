import numpy as np
import pickle
import tensorflow as tf
import matplotlib.pyplot as plt
import pandas as pd
import openpyxl
from openpyxl import load_workbook

config = tf.ConfigProto()
config.gpu_options.allow_growth=True
tf.debugging.set_log_device_placement(True)

output_folder = "200222"
run_dir = "../output/"+output_folder+"/models"

cnn_results = pd.read_csv("../output/"+output_folder+"/cnn_performance_table.csv")
best_val = cnn_results['validation'].iloc[0]

# indices of variables to modify in scenario testing
# industrial coal, transportation, residential coal
# INDICES NEED TO CHANGE IN THE UPDATED DATASET
'''
# 191206
scenario_variables = [3, 9, 10]
variables_export = ['AGC', 'AGN', 'AGO', 'INDCT', 'INDNT',
                    'INDOT', 'SVC', 'SVN', 'SVO', 'trans',
                    'RDC', 'RDN', 'DEM', 'rain', 'TEM']
# require gradients to be positive: AGO, INDCT, SVC, SVO, TRANS, RDC 
major_variables = [2,3,6,8,9,10]
oil_coal_variables = [0,2,3,5,6,8,9,10]
gas_variables = [1,4,7,11]

# 191219
scenario_variables = [2,6,7]
variables_export = ['AGC', 'AGO', 'INDCT',
                    'INDOT', 'SVC', 'SVO', 'trans',
                    'RDC', 'DEM', 'rain', 'TEM']
# require gradients to be positive: AGO, INDCT, SVC, SVO, TRANS, RDC
major_variables = [1,2,4,5,6,7]

#200111
scenario_variables = [0, 2, 6]
variables_export = ['COAL', 'AGO', 'INDCT',
                    'INDOT', 'SVC', 'SVO', 'trans',
                    'DEM', 'rain', 'TEM']
# require gradients to be positive: AGO, INDCT, SVC, SVO, TRANS, COAL (RDC+AGC)
major_variables = [0,1,2,4,5,6]
'''

#200222 and 200317
scenario_variables = [0,2,5]
variables_export = ['COAL', 'AGO', 'INDCT',
                    'INDOT', 'SVC', 'SVOtrans',
                    'DEM', 'rain', 'TEM']
# require gradients to be positive: AGO, INDCT, SVC, SVO, TRANS, COAL (RDC+AGC)
major_variables = [0,1,2,4,5]
indct = 2
coal = 0


# Inputs
with open('../data/project2_energy_1812/process/full_data_process_dic_no_airport_no_gas_coal_combined_no_SVO.pickle', 'rb') as data_standard:
    data_full_package = pickle.load(data_standard)

cnn_data_name = 'energy_norm_air_nonstand'
input_cnn_training = data_full_package[cnn_data_name]['input_training']
input_cnn_validation = data_full_package[cnn_data_name]['input_validation']
input_cnn_testing = data_full_package[cnn_data_name]['input_testing']
output_cnn_training = data_full_package[cnn_data_name]['output_training'][:, 0][:, np.newaxis]
output_cnn_validation = data_full_package[cnn_data_name]['output_validation'][:, 0][:, np.newaxis]
output_cnn_testing = data_full_package[cnn_data_name]['output_testing'][:, 0][:, np.newaxis]
output_cnn_all_vars_training = data_full_package[cnn_data_name]['output_training']
output_cnn_all_vars_validation = data_full_package[cnn_data_name]['output_validation']
output_cnn_all_vars_testing = data_full_package[cnn_data_name]['output_testing']
index_cnn_training = data_full_package[cnn_data_name]['index_training']
index_cnn_validation = data_full_package[cnn_data_name]['index_validation']
index_cnn_testing = data_full_package[cnn_data_name]['index_testing']
weights_cnn_training = data_full_package[cnn_data_name]['weight_training'].T/100
weights_cnn_validation = data_full_package[cnn_data_name]['weight_validation'].T/100
weights_cnn_testing = data_full_package[cnn_data_name]['weight_testing'].T/100
mean_data_name = 'energy_mean_air_nonstand'
input_mean_training = data_full_package[mean_data_name]['input_training']
input_mean_validation = data_full_package[mean_data_name]['input_validation']
input_mean_testing = data_full_package[mean_data_name]['input_testing']
std_data_name = 'energy_std_air_nonstand'
input_std_training = data_full_package[std_data_name]['input_training']
input_std_validation = data_full_package[std_data_name]['input_validation']
input_std_testing = data_full_package[std_data_name]['input_testing']
'''raw_input_data_name = 'energy_raw_nonstand'
input_cnn_training_raw = data_full_package[raw_input_data_name]['input_training']
'''

train_images=input_cnn_training
train_y=output_cnn_training
train_weights = weights_cnn_training
validation_images= input_cnn_validation
validation_y =output_cnn_validation
validation_weights = weights_cnn_validation
test_images =input_cnn_testing
test_y =output_cnn_testing
test_weights = weights_cnn_testing

train_mean_images = input_mean_training / 10000
validation_mean_images = input_mean_validation  / 10000
test_mean_images = input_mean_testing / 10000
train_std_images = input_std_training  / 10000
validation_std_images = input_std_validation / 10000
test_std_images = input_std_testing / 10000

all_trials_scenario_output = {}
all_trials_tr_gradients = {}
all_trials_val_gradients = {}
all_trials_te_gradients = {}
all_trials_tr_gradients_mean = {}
all_trials_val_gradients_mean = {}
all_trials_te_gradients_mean = {}

valid_models = []
model_validity = []
rank = 0
top_models = [57,308,165]

for index, val in zip(cnn_results['index'].astype(int), cnn_results['validation']):

    # if prediction error too large, then terminate analysis
    if val > best_val * 1.5:
        print('Results 50% worse than best model. Terminating...')
        break

    tf.reset_default_graph()

    sess = tf.Session(config=config)
    saver = tf.train.import_meta_graph(run_dir + "/model_" + str(index) + ".ckpt.meta")
    saver.restore(sess, run_dir + '/model_' + str(index) + ".ckpt")
    graph = tf.get_default_graph()

    X = graph.get_tensor_by_name("inputs/X:0")
    X_mean = graph.get_tensor_by_name("inputs/X_mean:0")
    y = graph.get_tensor_by_name("inputs/y:0")
    output = graph.get_tensor_by_name("outputs/output:0")
    weights = graph.get_tensor_by_name("inputs/weights:0")
    bias = graph.get_tensor_by_name("outputs/bias:0")

    # gradients w.r.t. images
    gradients = tf.gradients(output, X)
    gradients_mean = tf.gradients(output, X_mean)

    # gradients w.r.t. mean
    tr_gradients_mean = sess.run(gradients_mean, feed_dict={X: train_images, X_mean: train_mean_images})[0]
    val_gradients_mean = sess.run(gradients_mean, feed_dict={X: validation_images, X_mean: validation_mean_images})[0]
    te_gradients_mean = sess.run(gradients_mean, feed_dict={X: test_images, X_mean: test_mean_images})[0]

    gradients_mean_mean = np.mean(np.concatenate((tr_gradients_mean, val_gradients_mean, te_gradients_mean), axis=0), axis=0)
    gradients_mean_std = np.std(np.concatenate((tr_gradients_mean, val_gradients_mean, te_gradients_mean), axis=0), axis=0)
    gradients_mean_pos_cnt = np.mean(np.concatenate((tr_gradients_mean, val_gradients_mean, te_gradients_mean), axis=0) > 0, axis=0)

    if top_models is not None:
        if index in top_models:
            pd.DataFrame(np.concatenate((tr_gradients_mean, val_gradients_mean, te_gradients_mean), axis=0), \
                     columns=variables_export).to_csv("../output/"+output_folder+"/model_gradient_mean_"+str(index)+".csv")
    elif rank < 3:
        pd.DataFrame(np.concatenate((tr_gradients_mean, val_gradients_mean, te_gradients_mean), axis=0), \
                 columns=variables_export).to_csv("../output/"+output_folder+"/model_gradient_mean_"+str(index)+".csv")
        rank += 1

    # replace 0 values with inf so that the gradients will be 0
    input_std_training[input_std_training == 0] = np.inf
    input_std_validation[input_std_validation == 0] = np.inf
    input_std_testing[input_std_testing == 0] = np.inf

    # This is the most 'correct' version of the gradient
    # not used since it probably creates more trouble than what it's worth due to small values
    # tr_gradients = [[(np.divide(1, input_std_training) * (61*61-1) / (61*61) - np.power(input_std_training, -3)/(61*61)*np.power(input_cnn_training_raw[:, i, j, :] - input_mean_training, 2)) * tr_gradients[0][:, i, j, :] for i in range(61)] for j in range(61)]

    # gradients w.r.t. raw image values (chain rule)
    tr_gradients = sess.run(gradients, feed_dict={X: train_images, X_mean: train_mean_images})
    val_gradients = sess.run(gradients, feed_dict={X: validation_images, X_mean: validation_mean_images})
    te_gradients = sess.run(gradients, feed_dict={X: test_images, X_mean: test_mean_images})
    tr_gradients = np.array([[np.divide(tr_gradients[0][:, i, j, :], input_std_training)  * 10000 + np.array(
        tr_gradients_mean) / (61 * 61) for i in range(61)] for j in range(61)])
    val_gradients = np.array([[np.divide(val_gradients[0][:, i, j, :], input_std_validation)  * 10000 + np.array(
        val_gradients_mean) / (61 * 61) for i in range(61)] for j in range(61)])
    te_gradients = np.array([[np.divide(te_gradients[0][:, i, j, :], input_std_testing)  * 10000 + np.array(
        te_gradients_mean) / (61 * 61) for i in range(61)] for j in range(61)])

    gradients_cell_mean = np.sum(np.concatenate((tr_gradients, val_gradients, te_gradients), axis=2), axis=(0,1))

    # Filter models based on validity of gradients
    gradients_cell_mean_mean = np.mean(gradients_cell_mean, axis=0)
    gradients_cell_mean_major = gradients_cell_mean_mean[[major_variables]]
    notvalid = (gradients_cell_mean_mean[coal] < gradients_cell_mean_mean[indct])
    '''
    # gas variables taken out of the model 191219
    notvalid = 0
    for gas_var in gas_variables:
        notvalid += np.sum(te_gradients_mean_mean[gas_var] > te_gradients_mean_mean[[oil_coal_variables]])
    '''
    bias = sess.run(bias)[0]
    model_validity.append([index, val, int(np.sum(gradients_cell_mean_major < 0) != 0), int(notvalid), bias] +\
                          gradients_cell_mean_mean.tolist() + gradients_mean_mean.tolist() + \
                          [gradients_mean_pos_cnt[v] for v in major_variables])
    sess.close()
    print(index, val)

'''
book = openpyxl.Workbook()
book.save("../output/"+output_folder+"/gradient_validity.xlsx")
book = load_workbook("../output/"+output_folder+"/gradient_validity.xlsx")
writer = pd.ExcelWriter("../output/"+output_folder+"/gradient_validity.xlsx", engine='openpyxl')
writer.book = book
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
for s in range(3):
    df3 = pd.DataFrame(data=scenario_output[s,:,:,0].T, columns = ['actual','fitted']+[str(int(pct)) + '%' for pct in np.arange(2,22,2)])
    df3.to_excel(writer, sheet_name=scenario_name[s], index=False,header=True)
book.remove(book['Sheet'])
writer.save()
'''

model_validity = pd.DataFrame(model_validity, columns=['index', 'validation', 'neg_gradients', 'coals', 'bias'] + \
                        [v + "cell_mean" for v in variables_export] + [v + "mean_mean" for v in variables_export] + \
                        [variables_export[v]+"pos_cnt" for v in major_variables]).to_csv("../output/"+output_folder+\
                        "/model_validity.csv", index=False)
