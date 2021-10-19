# -*- coding: utf-8 -*-
"""
Created on Tue Oct  8 22:32:32 2019

@author: qingyi
"""

import glob
import pickle as pkl
import numpy as np
import pickle
import tensorflow as tf
import matplotlib.pyplot as plt
import pandas as pd
import sys
import openpyxl
from openpyxl import load_workbook
import os

from setup import *
import util_data

#os.environ["CUDA_VISIBLE_DEVICES"]="-1"
os.environ['TF_CPP_MIN_LOG_LEVEL'] = '3'
config = tf.ConfigProto()
config.gpu_options.allow_growth = True
tf.debugging.set_log_device_placement(True)

output_var = [0]
radius = 30
output_folder = "211010"
top_models = [9]

hp_idx = 145
linear_coef = 0.3
run_suffix = '_linearfix'

run_dir = output_dir+output_folder+"/models/models_" + "".join([str(ov) for ov in output_var]) + "_" + str(radius)
char_name = '_no_airport_no_gas_coal_combined_oil_combined'
standard = 'const'

# indices of variables to modify in scenario testing
# industrial coal, transportation, residential coal
# INDICES NEED TO CHANGE IN THE UPDATED DATASET
'''
# 191206
scenario_variables = [3, 9, 10]
variables_export = ['AGC', 'AGN', 'AGO', 'INDCT', 'INDNT',
                    'INDOT', 'SVC', 'SVN', 'SVO', 'trans',
                    'RDC', 'RDN', 'DEM', 'rain', 'TEM']
# 191219
scenario_variables = [2,6,7]
variables_export = ['AGC', 'AGO', 'INDCT',
                    'INDOT', 'SVC', 'SVO', 'trans',
                    'RDC', 'DEM', 'rain', 'TEM']
#200111
scenario_variables = [0, 2, 6]
variables_export = ['COAL', 'AGO', 'INDCT',
                    'INDOT', 'SVC', 'SVO', 'trans',
                    'DEM', 'rain', 'TEM']
#200222 and 200317 and 200417
scenario_variables = [0,2,5]
scenario_name = ['res_col', 'ind_col', 'trn_oil']
variables_export = ['COAL', 'AGO', 'INDCT',
                    'INDOT', 'SVC', 'SVOtrans',
                    'DEM', 'rain', 'TEM']'''
#20060601/200905
scenario_variables = [0,1,4]
scenario_name = ['res_col', 'ind_col', 'trn_oil']
variables_export = ['COAL', 'INDCT',
                    'INDOT', 'SVC', 'OIL',
                    'DEM', 'rain', 'TEM']

# require gradients to be positive: AGO, INDCT, SVC, SVO, TRANS, COAL (RDC+AGC)
major_variables = [0,1,2,3,4]

# Inputs
station_index = pd.read_excel(data_dir+'raw/stationID.xlsx')
(train_images,train_y,train_weights,
    validation_images,validation_y,validation_weights,
    test_images,test_y,test_weights,
    train_mean_images,validation_mean_images,test_mean_images,
    index_cnn_training,index_cnn_validation,index_cnn_testing,
    sector_max) = util_data.read_data(char_name, standard, radius, output_var)

control_var_training, control_var_validation, control_var_testing, control_scale = \
    util_data.get_control_variables(filename='agriculture_variables_station.xlsx',
                                train_index=index_cnn_training,
                                validation_index=index_cnn_validation,
                                test_index=index_cnn_testing)

with open(data_dir+'process/energy_data_dic' + char_name + '.pickle', 'rb') as data_dic:
    data_raw = pickle.load(data_dic)

# a list of station codes (string)
indices = np.concatenate((index_cnn_training, index_cnn_validation, index_cnn_testing))

batch_size = 128

# get a mask - whether the original emission is 0
# get the total amount of emissions by sector
id10 = []
mask = {}
emissions = {}
for v in variables_export:
    mask[v] = []
    emissions[v] = []
for s in range(943):
    folder = station_index[station_index['station_code'] == indices[s]]['STID'].iloc[0]
    pixelID = pd.read_csv(data_dir+'raw/result/' + str(folder) + '/ID10.csv', header=None).to_numpy()
    id10 += pixelID.flatten().tolist()
    for v in variables_export:
        mask[v] += (data_raw[indices[s]][v].values.flatten() != 0).tolist()
        emissions[v] += data_raw[indices[s]][v].values.tolist()

# turn the mask into a dataframe - one cell might belong to multiple stations
# mask_df: group by id10
# mask_full_df: sequential, not grouped
# emissions_df: groupby id10 and take mean so that only one cell value is counted towards the total
id10 = np.array(id10).reshape((-1, 1))
mask_df = id10
emissions_df = id10
for v in variables_export:
    mask[v] = np.array(mask[v]).reshape((-1, 1))
    emissions[v] = np.array(emissions[v]).reshape((-1, 1))
    mask_df = np.append(mask_df, mask[v], axis=1)
    emissions_df = np.append(emissions_df, emissions[v], axis=1)
mask_df = pd.DataFrame(mask_df, index=id10.flatten(), columns = ['id10'] + variables_export)
emissions_df = pd.DataFrame(emissions_df, index=id10.flatten(), columns = ['id10'] + variables_export)
emissions_df = emissions_df.groupby('id10').mean()
mask_full_df = mask_df.copy()
mask_df = mask_df.groupby('id10', as_index=False).max()
# total emissions by sector
print("Total Emissions:", emissions_df.sum())

for model_idx in top_models:
    print("Model: ", model_idx)
    print('Calculating gradients...')

    tf.reset_default_graph()

    if linear_coef == 0:
        linear_coef = int(linear_coef)
    files = glob.glob(run_dir + "/model_"+str(linear_coef) +"_"+str(hp_idx) + "_" +str(model_idx)+run_suffix+"_*.ckpt.meta")
    if len(files) == 0:
        print("Model not found. Exiting...")
        print(run_dir + "/model_"+str(linear_coef) +"_"+str(hp_idx) + "_" +str(model_idx)+run_suffix+"_*.ckpt.meta")
        sys.exit()
    elif len(files) > 1:
        print('More than one model found. Exiting...')
        sys.exit()
    else:
        file = files[0]

    sess = tf.Session(config=config)
    saver = tf.train.import_meta_graph(file)
    saver.restore(sess, '.'.join(file.split('.')[:-1]))
    graph = tf.get_default_graph()

    X = graph.get_tensor_by_name("inputs/X:0")
    X_mean = graph.get_tensor_by_name("inputs/X_mean:0")
    y = graph.get_tensor_by_name("inputs/y:0")
    output = graph.get_tensor_by_name("outputs/final_output:0")
    weights = graph.get_tensor_by_name("inputs/weights:0")

    # gradients w.r.t. images
    gradients = tf.gradients(output, X)

    # op=graph.get_operations()
    # print([m.values() for m in op])
    #
    # break
    # scenario testing

    print('Scenario testing...')
    scenario_output = []
    for variable in scenario_variables:
        temp = [list(train_y)+list(validation_y)+list(test_y)]
        for factor in np.arange(0, 0.22, 0.02).tolist():
            train_images_update = np.copy(train_images)
            train_images_update[:,:,:,variable] = train_images_update[:,:,:, variable] * (1 - factor)
            train_mean_update = np.copy(np.hstack([train_mean_images, control_var_training]))
            train_mean_update[:,variable] = train_mean_update[:,variable] * (1-factor)
            y_train_scenario = sess.run(output, feed_dict={X: train_images_update, X_mean: train_mean_update})

            val_images_update = np.copy(validation_images)
            val_images_update[:,:,:, variable] = val_images_update[:,:,:, variable] * (1 - factor)
            val_mean_update = np.copy(np.hstack([validation_mean_images, control_var_validation]))
            val_mean_update[:, variable] = val_mean_update[:,variable] * (1 - factor)
            y_val_scenario = sess.run(output, feed_dict={X: val_images_update, X_mean: val_mean_update})

            test_images_update = np.copy(test_images)
            test_images_update[:,:,:, variable] = test_images_update[:,:,:, variable] * (1 - factor)
            test_mean_update = np.copy(np.hstack([test_mean_images, control_var_testing]))
            test_mean_update[:, variable] = test_mean_update[:,variable] * (1 - factor)
            y_test_scenario = sess.run(output, feed_dict={X: test_images_update, X_mean: test_mean_update})

            temp.append(list(y_train_scenario) + list(y_val_scenario) + list(y_test_scenario))
            if factor == 0:
                fitted_y_train = y_train_scenario
                fitted_y_val = y_val_scenario
                fitted_y_test = y_test_scenario
        scenario_output.append(temp)

    scenario_output = np.array(scenario_output)
    book = openpyxl.Workbook()
    book.save(output_dir+output_folder+"/"+str(hp_idx)+"_"+str(linear_coef)+"_"+str(model_idx)+"_scenario_pm"+run_suffix+".xlsx")
    book = load_workbook(output_dir+output_folder+"/"+str(hp_idx)+"_"+str(linear_coef)+"_"+str(model_idx)+"_scenario_pm"+run_suffix+".xlsx")
    writer = pd.ExcelWriter(output_dir+output_folder+"/"+str(hp_idx)+"_"+str(linear_coef)+"_"+str(model_idx)+"_scenario_pm"+run_suffix+".xlsx", engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    for s in range(3):
        df3 = pd.DataFrame(data=scenario_output[s,:,:,0].T, columns = ['actual','fitted']+[str(int(pct)) + '%' for pct in np.arange(2,22,2)])
        df3.to_excel(writer, sheet_name=scenario_name[s], index=False,header=True)
    book.remove(book['Sheet'])
    writer.save()

    # PM under different fractions of emissions
    '''
    temp = []
    factor_list = [0,5,10,20,30,40,50,60,70,80,90,95,99]
    for factor in factor_list:
        frac = factor / 100
        train_images_update = np.copy(train_images)
        train_images_update = train_images_update * frac
        y_train_scenario = sess.run(output, feed_dict={X: train_images_update})
        val_images_update = np.copy(validation_images)
        val_images_update = val_images_update * frac
        y_val_scenario = sess.run(output, feed_dict={X: val_images_update})
        test_images_update = np.copy(test_images)
        test_images_update = test_images_update * frac
        y_test_scenario = sess.run(output, feed_dict={X: test_images_update})
        fitted = list(y_train_scenario) + list(y_val_scenario) + list(y_test_scenario)
        weighted_pm = np.average(np.array(fitted).flatten(),
                                 weights=np.concatenate((train_weights, validation_weights, test_weights)).flatten())
        temp.append(weighted_pm)
    weighted_pm = np.average(np.array(list(fitted_y_train) + list(fitted_y_val) + list(fitted_y_test)).flatten(),
                             weights=np.concatenate((train_weights, validation_weights, test_weights)).flatten())
    temp.append(weighted_pm)
    fig, ax = plt.subplots(figsize=(6,4))
    ax.plot(factor_list + [100], temp)
    ax.plot([0, 100], [temp[0], temp[-1]], '--', label='Average Sensitivity')
    ax.plot([0, 100], [temp[-1] - 100*(temp[-1] - temp[-2]), temp[-1]], '--', label='Marginal Sensitivity')
    ax.legend()
    ax.set_xlabel("Frac of emissions, %")
    ax.set_ylabel("Weighted PM 2.5")
    fig.savefig(output_dir + output_folder + "/plots/"+str(hp_idx)+"_"+str(linear_coef)+"_"+str(model_idx)+"_marginal_avg_sensitivity"+run_suffix+".png", bbox_inches='tight')
    '''

    # gradients w.r.t. raw image values (chain rule）
    tr_gradients = sess.run(gradients, feed_dict={X: train_images, X_mean: np.hstack([train_mean_images, control_var_training])})
    tr_gradients = 10000*np.array([[[np.divide(tr_gradients[0][:, i, j, l], sector_max[k]) for k,l in zip(variables_export, range(len(variables_export)))] for i in range(61)] for j in range(61)])

    val_gradients = sess.run(gradients, feed_dict={X: validation_images, X_mean: np.hstack([validation_mean_images, control_var_validation])})
    val_gradients = 10000*np.array([[[np.divide(val_gradients[0][:, i, j, l], sector_max[k]) for k,l in zip(variables_export, range(len(variables_export)))] for i in range(61)] for j in range(61)])

    te_gradients = sess.run(gradients, feed_dict={X: test_images, X_mean: np.hstack([test_mean_images, control_var_testing])})
    te_gradients = 10000*np.array([[[np.divide(te_gradients[0][:, i, j, l], sector_max[k]) for k,l in zip(variables_export, range(len(variables_export)))] for i in range(61)] for j in range(61)])

    # #sector x #observations
    tr_gradients_mean = np.sum(tr_gradients, axis=(0,1))
    val_gradients_mean = np.sum(val_gradients, axis=(0,1))
    te_gradients_mean = np.sum(te_gradients, axis=(0,1))

    with open(output_dir + output_folder + "/"+str(hp_idx)+"_"+str(linear_coef)+"_"+str(model_idx)+"_gradients_by_cell"+run_suffix+".pkl", "wb") as f:
        pkl.dump(tr_gradients, f)
        pkl.dump(val_gradients, f)
        pkl.dump(te_gradients, f)
        pkl.dump(tr_gradients_mean, f)
        pkl.dump(val_gradients_mean, f)
        pkl.dump(te_gradients_mean, f)

    print('Fill in gradients...')

    gradients_all = np.concatenate((tr_gradients, val_gradients, te_gradients), axis=3)
    gradients = None
    for i in range(gradients_all.shape[3]):
    # iterate through stations and flatten gradients
        if gradients is None:
            gradients = np.reshape(gradients_all[:,:,:,i], (-1, gradients_all.shape[2]))
        else:
            gradients = np.concatenate((gradients, np.reshape(gradients_all[:,:,:,i], (-1, gradients_all.shape[2]))), axis=0)

    gradients_full = pd.DataFrame(np.append(id10, gradients, axis=1), index=id10.flatten(), columns = ['id10'] + variables_export)
    gradients = gradients_full.groupby('id10', as_index=False).max()

    gradients_masked = gradients * mask_df
    gradients_masked['id10'] = np.sqrt(gradients_masked['id10'])

    '''
    # not used
    gradients_full_masked = gradients_full * mask_full_df
    gradients_full_masked['id10'] = np.sqrt(gradients_full_masked['id10'])
    '''

    with open(output_dir+output_folder+"/"+str(hp_idx)+"_"+str(linear_coef)+"_"+str(model_idx)+"_results"+run_suffix+".pkl", "wb") as f:
        pkl.dump(scenario_output, f)
        pkl.dump(gradients, f)
        pkl.dump(gradients_masked, f)
        pkl.dump(gradients_full, f)
        pkl.dump(mask_full_df, f)
        pkl.dump(id10, f)
        pkl.dump(indices, f)
        pkl.dump(fitted_y_train, f)
        pkl.dump(fitted_y_val, f)
        pkl.dump(fitted_y_test, f)

    gradients.to_csv(output_dir+output_folder+"/"+str(hp_idx)+"_"+str(linear_coef)+"_"+str(model_idx)+"_gradients"+run_suffix+".csv", index=False)
    gradients_masked.to_csv(output_dir+output_folder+"/"+str(hp_idx)+"_"+str(linear_coef)+"_"+str(model_idx)+"_gradients_masked"+run_suffix+".csv", index=False)


df1 = pd.DataFrame({'Station indices': indices})
df1.to_excel(output_dir+output_folder+"/station_indices"+run_suffix+".xlsx", sheet_name='sheet1', index=False,header=False)

'''
# testing the limits of the stationIDs
max_id = 0
min_id = 10000
for folder in listdir('../data/project2_energy_1812/raw/result'):
    if 'DS' not in folder and '._' not in folder: # address the DS_Store...
        for csv_file in listdir('../data/project2_energy_1812/raw/result/'+ folder):
            directory_path = '../data/project2_energy_1812/raw/result/'+folder+'/ID10.csv'
            pixelID = pd.read_csv(directory_path, header = None)
            pixelID = pixelID[((pixelID != 9999999.0) & (pixelID != 8888888.0))]
            if np.max(np.max(pixelID)) > max_id:
                max_id = np.max(np.max(pixelID))
                print(max_id)
            if np.min(np.min(pixelID)) < min_id:
                min_id = np.min(np.min(pixelID))
                print(min_id)
'''
