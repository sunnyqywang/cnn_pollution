from GEMM_function_20200328 import cal
from GEMM_function_mean_only_0328 import cal_mean
import numpy as np
import openpyxl
from openpyxl import load_workbook
import pandas as pd
import pickle as pkl
import time
import sys

output_folder = '200905'
top_models = [72,87,40,122,7]
variables_export = ['COAL', 'INDCT',
                    'INDOT', 'SVC', 'OIL',
                    'DEM', 'rain', 'TEM']
major_variables = [0,1,2,3,4]

char_name = '_no_airport_no_gas_coal_combined_oil_combined'
standard = 'const'

with open('../data/project2_energy_1812/process/full_data_process_dic'+ char_name + '.pickle', 'rb') as data_standard:
    data_full_package = pkl.load(data_standard)

cnn_data_name = 'energy_'+standard+'_air_nonstand'
output_cnn_training = data_full_package[cnn_data_name]['output_training'][:, 0]
output_cnn_validation = data_full_package[cnn_data_name]['output_validation'][:, 0]
output_cnn_testing = data_full_package[cnn_data_name]['output_testing'][:, 0]
weights_cnn_training = data_full_package[cnn_data_name]['weight_training'].T
weights_cnn_validation = data_full_package[cnn_data_name]['weight_validation'].T
weights_cnn_testing = data_full_package[cnn_data_name]['weight_testing'].T
index_cnn_training = data_full_package[cnn_data_name]['index_training']
index_cnn_validation = data_full_package[cnn_data_name]['index_validation']
index_cnn_testing = data_full_package[cnn_data_name]['index_testing']

#pm0 = np.concatenate((output_cnn_training, output_cnn_validation, output_cnn_testing))
population = np.concatenate((weights_cnn_training, weights_cnn_validation, weights_cnn_testing))[:, 0]
indices = np.concatenate((index_cnn_training, index_cnn_validation, index_cnn_testing))

for index in top_models:

    with open("../output/"+output_folder+"/"+str(index)+"_results.pkl", "rb") as f:
        scenario_output = np.array(pkl.load(f))[:, :, :, 0]
        pkl.load(f)  #gradients
        pkl.load(f) #gradients_masked
        gradients_full = pkl.load(f)
        mask_full_df = pkl.load(f)
        id10_full = pkl.load(f)
        indices = pkl.load(f)

    mask_full_df['id10_all_true'] = True

    # marginal avoidance based on gradients
    grid_size = 61*61
    mean_df_full = pd.DataFrame(id10_full, columns=['id10'])
    # the original PM value: 0th scenario variable (can be 1 or 2 as well, doesn't matter),
    # the second column (first being the actual value), all stations
    pm0 = scenario_output[0,1,:]
    for vi in major_variables:
        v = variables_export[vi]
        print(v)
        mean = []
        p0 = pm0
        pop = population
        gr = gradients_full[v].to_numpy()
        mask = mask_full_df[v].to_numpy(dtype=bool)

        for i in range(len(gr)):
            p1 = p0[i//grid_size-1]-gr[i]
            dy = cal_mean(p0[i // grid_size - 1], p1, pop[i // grid_size - 1])
            mean.append(dy[0])

        mean_df_full[v] = -np.array(mean) * 1.8 * np.power(10, 10)

    mean_df_full.to_csv("../output/"+output_folder+"/"+str(index)+"_md_mean_full.csv", index=False)
    mean_df = mean_df_full.groupby('id10').sum()
    mean_df.to_csv("../output/"+output_folder+"/"+str(index)+"_md_mean.csv")

    # scenario
    mean = []
    p5 = []
    p95 = []
    count = 0
    result = []
    for v in range(scenario_output.shape[0]): # variable
        pm0 = scenario_output[v,1,:]
        for sce in range(1, scenario_output.shape[1]): # scenario numbers
            for p0, p1, pop, station_index in zip(pm0, scenario_output[v,sce,:], population, indices):
                dy = cal(p0, p1, pop)
                # should be negative if pm1 less than pm0
                result.append([v, sce, station_index, p0, p1, pop, dy[0], dy[1], dy[2]])
                if p0>p1 and dy[0] > 0 :
                    print(v, sce, p0, p1, pop, dy)
                    sys.exit()
                mean.append(dy[0])
                p5.append(dy[1])
                p95.append(dy[2])

    output_shape = list(scenario_output.shape)
    output_shape[1] -= 1
    scenario_pd_mean = np.reshape(mean, output_shape)
    scenario_pd_p5 = np.reshape(p5, output_shape)
    scenario_pd_p95 = np.reshape(p95, output_shape)

    book = openpyxl.Workbook()
    book.save("../output/" + output_folder + '/'+str(index)+'_scenario_ad_xlsx.xlsx')
    book = load_workbook("../output/" + output_folder + '/'+str(index)+'_scenario_ad_xlsx.xlsx')
    writer = pd.ExcelWriter("../output/" + output_folder + '/'+str(index)+'_scenario_ad_xlsx.xlsx', engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    scenario_name = ['res_col', 'ind_col', 'trn_oil']
    for s in range(3):
        print(scenario_name[s])
        print('Mean Avoided Death: (20% reduction)')
        print(scenario_pd_mean[s,-1,:].sum())
        print('p5 Avoided Death: (20% reduction)')
        print(scenario_pd_p5[s,-1,:].sum())
        print('p95 Avoided Death: (20% reduction)')
        print(scenario_pd_p95[s,-1,:].sum())

        df1 = pd.DataFrame(data=scenario_pd_mean[s, :, :].T)
        df2 = pd.DataFrame(data=scenario_pd_p5[s, :, :].T)
        df3 = pd.DataFrame(data=scenario_pd_p95[s, :, :].T)
        df1.to_excel(writer, sheet_name=scenario_name[s]+'_mean', index=False, header=False)
        df2.to_excel(writer, sheet_name=scenario_name[s]+'_p5', index=False, header=False)
        df3.to_excel(writer, sheet_name=scenario_name[s]+'_p95', index=False, header=False)
    book.remove(book['Sheet'])
    writer.save()

    result = pd.DataFrame(result, columns=['variable','scenario','index','p0','p1','pop','mean','p5','p95']).to_csv("../output/" \
                        + output_folder + '/'+str(index)+'_scenario_ad_csv.csv', index=False, float_format='%.3f')



'''
# vector implementation (old version)
mean = np.array([])
p5 = np.array([])
p95 = np.array([])
grid_size = 61*61
step = 61*61
start_time = time.time()

for v in variables_export:
    print(v)
    #p0 = np.repeat(pm0, grid_size)
    #pop = np.repeat(population, grid_size)
    p0 = pm0
    pop=population
    gr = gradients_full[v].to_numpy()
    mask = mask_full_df[v].to_numpy(dtype=bool)
    for i in np.arange(step, len(gr)+step, step):
        print(i)
        if i > len(gradients_full):
            p1 = p0[i//grid_size-1]-gr[i-step:]
            mask1 = mask[i-step:]
            dy = cal_premature_death_vec(p0[i//step-1], p1[mask1], pop[i//step-1])
        else:
            p1 = p0[i//grid_size-1]-gr[i-step:i]
            mask1 = mask[i-step:i]
            dy = cal_premature_death_vec(p0[i//step-1], p1[mask1], pop[i//step-1])

    mean = np.concatenate((mean, dy[0]))
    p5 = np.concatenate((p5, dy[1]))
    p95 = np.concatenate((p95, dy[2]))
    print(time.time() - start_time)

mean = np.array(mean).reshape((len(variables_export), -1)).T
mean = pd.DataFrame(np.append(id10_full[mask], mean, axis=1), index=id10_full[mask], columns = ['id10'] + variables_export)
mean = mean.groupby('id10', as_index=False).sum()
mean.to_csv("../output/"+output_folder+"/pd_mean_"+str(index)+".csv", index=False)

p5 = np.array(p5).reshape((len(variables_export), -1)).T
p5 = pd.DataFrame(np.append(id10_full[mask], p5, axis=1), index=id10_full[mask], columns = ['id10'] + variables_export)
p5 = p5.groupby('id10', as_index=False).sum()
p5.to_csv("../output/"+output_folder+"/pd_p5_"+str(index)+".csv", index=False)

p95 = np.array(p95).reshape((len(variables_export), -1)).T
p95 = pd.DataFrame(np.append(id10_full[mask], p95, axis=1), index=id10_full[mask], columns = ['id10'] + variables_export)
p95 = p95.groupby('id10', as_index=False).sum()
p95.to_csv("../output/"+output_folder+"/pd_p95_"+str(index)+".csv", index=False)
'''
