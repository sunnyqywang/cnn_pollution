# -*- coding: utf-8 -*-
"""
Created on Tue Oct  8 22:32:32 2019

@author: wangqi44
"""

import pickle as pkl
import numpy as np
import pickle
import tensorflow as tf
import matplotlib.pyplot as plt
import pandas as pd
import sys
import openpyxl
from openpyxl import load_workbook
import os
os.environ["CUDA_VISIBLE_DEVICES"]="-1"

output_folder = "200317"
top_models = [9,7,0]
config = tf.ConfigProto()
#config.gpu_options.allow_growth = True
#tf.debugging.set_log_device_placement(True)

run_dir = "../output/" + output_folder + "/models"
char_name = '_no_airport_no_gas_coal_combined_no_SVO'

# indices of variables to modify in scenario testing
# industrial coal, transportation, residential coal
# INDICES NEED TO CHANGE IN THE UPDATED DATASET
'''
# 191206
scenario_variables = [3, 9, 10]
variables_export = ['AGC', 'AGN', 'AGO', 'INDCT', 'INDNT',
                    'INDOT', 'SVC', 'SVN', 'SVO', 'trans',
                    'RDC', 'RDN', 'DEM', 'rain', 'TEM']
# 191219
scenario_variables = [2,6,7]
variables_export = ['AGC', 'AGO', 'INDCT',
                    'INDOT', 'SVC', 'SVO', 'trans',
                    'RDC', 'DEM', 'rain', 'TEM']
#200111
scenario_variables = [0, 2, 6]
variables_export = ['COAL', 'AGO', 'INDCT',
                    'INDOT', 'SVC', 'SVO', 'trans',
                    'DEM', 'rain', 'TEM']'''
#200222 and 200317
scenario_variables = [0,2,5]
scenario_name = ['res_col', 'ind_col', 'trn_oil']
variables_export = ['COAL', 'AGO', 'INDCT',
                    'INDOT', 'SVC', 'SVOtrans',
                    'DEM', 'rain', 'TEM']
# require gradients to be positive: AGO, INDCT, SVC, SVO, TRANS, COAL (RDC+AGC)
major_variables = [0,1,2,4,5]

# Input
station_index = pd.read_excel('../data/project2_energy_1812/raw/stationID.xlsx')
with open('../data/project2_energy_1812/process/full_data_process_dic' + char_name + '.pickle', 'rb') as data_standard:
    data_full_package = pickle.load(data_standard)

cnn_data_name = 'energy_minmax_air_nonstand'
input_cnn_training = data_full_package[cnn_data_name]['input_training']
input_cnn_validation = data_full_package[cnn_data_name]['input_validation']
input_cnn_testing = data_full_package[cnn_data_name]['input_testing']
output_cnn_training = data_full_package[cnn_data_name]['output_training'][:, 0][:, np.newaxis]
output_cnn_validation = data_full_package[cnn_data_name]['output_validation'][:, 0][:, np.newaxis]
output_cnn_testing = data_full_package[cnn_data_name]['output_testing'][:, 0][:, np.newaxis]
output_cnn_all_vars_training = data_full_package[cnn_data_name]['output_training']
output_cnn_all_vars_validation = data_full_package[cnn_data_name]['output_validation']
output_cnn_all_vars_testing = data_full_package[cnn_data_name]['output_testing']
index_cnn_training = data_full_package[cnn_data_name]['index_training']
index_cnn_validation = data_full_package[cnn_data_name]['index_validation']
index_cnn_testing = data_full_package[cnn_data_name]['index_testing']
weights_cnn_training = data_full_package[cnn_data_name]['weight_training'].T / 100
weights_cnn_validation = data_full_package[cnn_data_name]['weight_validation'].T / 100
weights_cnn_testing = data_full_package[cnn_data_name]['weight_testing'].T / 100
min_data_name = 'energy_min_air_nonstand'
input_min_training = data_full_package[min_data_name]['input_training']
input_min_validation = data_full_package[min_data_name]['input_validation']
input_min_testing = data_full_package[min_data_name]['input_testing']
max_data_name = 'energy_max_air_nonstand'
input_max_training = data_full_package[max_data_name]['input_training']
input_max_validation = data_full_package[max_data_name]['input_validation']
input_max_testing = data_full_package[max_data_name]['input_testing']
train_max_images = input_max_training / 10000
validation_max_images = input_max_validation  / 10000
test_max_images = input_max_testing / 10000

train_images = input_cnn_training
train_y = output_cnn_training
train_weights = weights_cnn_training
validation_images = input_cnn_validation
validation_y = output_cnn_validation
validation_weights = weights_cnn_validation
test_images = input_cnn_testing
test_y = output_cnn_testing
test_weights = weights_cnn_testing

with open('../data/project2_energy_1812/process/energy_data_dic' + char_name + '.pickle', 'rb') as data_dic:
    data_raw = pickle.load(data_dic)

# a list of station codes (string)
indices = np.concatenate((index_cnn_training, index_cnn_validation, index_cnn_testing))

batch_size = 100

id10 = []
mask = {}
for v in variables_export:
    mask[v] = []
for s in range(943):
    folder = station_index[station_index['station_code'] == indices[s]]['STID'].iloc[0]
    pixelID = pd.read_csv('../data/project2_energy_1812/raw/result/' + str(folder) + '/ID10.csv',
                          header=None).to_numpy()
    id10 += pixelID.flatten().tolist()
    for v in variables_export:
        mask[v] += (data_raw[indices[s]][v].values.flatten() != 0).tolist()

id10 = np.array(id10).reshape((-1, 1))
mask_df = id10
for v in variables_export:
    mask[v] = np.array(mask[v]).reshape((-1, 1))
    mask_df = np.append(mask_df, mask[v], axis=1)
mask_df = pd.DataFrame(mask_df, index=id10, columns = ['id10'] + variables_export)
mask_full_df = mask_df.copy()
mask_df = mask_df.groupby('id10', as_index=False).max()

print(np.sum(mask_df == 0) / len(mask_df)/ 7)

for index in top_models:
    print("Model: ", index)
    print('Calculating gradients...')

    tf.reset_default_graph()

    sess = tf.Session(config=config)
    saver = tf.train.import_meta_graph(run_dir + "/model_" + str(index) + ".ckpt.meta")
    saver.restore(sess, run_dir + '/model_' + str(index) + ".ckpt")
    graph = tf.get_default_graph()

    X = graph.get_tensor_by_name("inputs/X:0")
    X_max = graph.get_tensor_by_name("inputs/X_max:0")
    y = graph.get_tensor_by_name("inputs/y:0")
    output = graph.get_tensor_by_name("outputs/output:0")
    weights = graph.get_tensor_by_name("inputs/weights:0")

    # gradients w.r.t. images
    gradients = tf.gradients(output, X)

    # scenario testing
    print('Scenario testing...')
    test_max_images[test_max_images == np.inf] = 0
    train_max_images[train_max_images == np.inf] = 0
    validation_max_images[validation_max_images == np.inf] = 0
    scenario_output = []
    for variable in scenario_variables:
        temp = [list(train_y)+list(validation_y)+list(test_y)]
        for factor in np.arange(0, 0.22, 0.02).tolist():
            train_max_update = np.copy(train_max_images)
            train_max_update[:, variable] = train_max_update[:, variable] * (1 - factor)
            y_train_scenario = sess.run(output, feed_dict={X: train_images, X_max: train_max_update})
            val_max_update = np.copy(validation_max_images)
            val_max_update[:, variable] = val_max_update[:, variable] * (1 - factor)
            y_val_scenario = sess.run(output, feed_dict={X: validation_images, X_max: val_max_update})
            test_max_update = np.copy(test_max_images)
            test_max_update[:, variable] = test_max_update[:, variable] * (1 - factor)
            y_test_scenario = sess.run(output, feed_dict={X: test_images, X_max: test_max_update})
            temp.append(list(y_train_scenario) + list(y_val_scenario) + list(y_test_scenario))
            if factor == 0:
                fitted_y_train = y_train_scenario
                fitted_y_val = y_val_scenario
                fitted_y_test = y_test_scenario
        scenario_output.append(temp)

    scenario_output = np.array(scenario_output)
    book = openpyxl.Workbook()
    book.save("../output/"+output_folder+"/scenario_output_"+str(index)+".xlsx")
    book = load_workbook("../output/"+output_folder+"/scenario_output_"+str(index)+".xlsx")
    writer = pd.ExcelWriter("../output/"+output_folder+"/scenario_output_"+str(index)+".xlsx", engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    for s in range(3):
        df3 = pd.DataFrame(data=scenario_output[s,:,:,0].T, columns = ['actual','fitted']+[str(int(pct)) + '%' for pct in np.arange(2,22,2)])
        df3.to_excel(writer, sheet_name=scenario_name[s], index=False,header=True)
    book.remove(book['Sheet'])
    writer.save()

    te_gradients = sess.run(gradients, feed_dict={X: test_images, X_max: test_max_images})
    test_max_images[test_max_images == 0] = np.inf
    te_gradients = np.array([[np.divide(te_gradients[0][:, i, j, :], test_max_images)  for i in range(61)] for j in range(61)])
    te_gradients_mean = np.sum(te_gradients, axis=(0,1))


    tr_gradients = sess.run(gradients, feed_dict={X: train_images, X_max: train_max_images})
    train_max_images[train_max_images == 0] = np.inf
    tr_gradients = np.array([[np.divide(tr_gradients[0][:, i, j, :], train_max_images)  for i in range(61)] for j in range(61)])
    tr_gradients_mean = np.sum(tr_gradients, axis=(0,1))


    val_gradients = sess.run(gradients, feed_dict={X: validation_images, X_max: validation_max_images})
    validation_max_images[validation_max_images == 0] = np.inf
    val_gradients = np.array([[np.divide(val_gradients[0][:, i, j, :], validation_max_images)  for i in range(61)] for j in range(61)])
    val_gradients_mean = np.sum(val_gradients, axis=(0,1))

    with open("../output/" + output_folder + "/gradients_by_cell_" + str(index) + ".pkl", "wb") as f:
        pkl.dump(tr_gradients, f)
        pkl.dump(val_gradients, f)
        pkl.dump(te_gradients, f)
        pkl.dump(tr_gradients_mean, f)
        pkl.dump(val_gradients_mean, f)
        pkl.dump(te_gradients_mean, f)

    print('Fill in gradients...')
    # Fill in gradients in a matrix form
    gradients_all = np.concatenate((tr_gradients, val_gradients, te_gradients), axis=2)
    gradients = None
    for i in range(gradients_all.shape[2]):
        if gradients is None:
            gradients = np.reshape(gradients_all[:,:,i,:], (-1, gradients_all.shape[3]))
        else:
            gradients = np.concatenate((gradients, np.reshape(gradients_all[:,:,i,:], (-1, gradients_all.shape[3]))), axis=0)

    gradients_full = pd.DataFrame(np.append(id10, gradients, axis=1), index=id10, columns = ['id10'] + variables_export)
    gradients = gradients_full.groupby('id10', as_index=False).max()

    gradients_masked = gradients * mask_df
    gradients_masked['id10'] = np.sqrt(gradients_masked['id10'])

    gradients_full_masked = gradients_full * mask_full_df
    gradients_full_masked['id10'] = np.sqrt(gradients_full_masked['id10'])

    with open("../output/"+output_folder+"/results_"+str(index)+".pkl", "wb") as f:
        pkl.dump(scenario_output, f)
        pkl.dump(gradients, f)
        pkl.dump(gradients_masked, f)
        pkl.dump(gradients_full, f)
        pkl.dump(mask_full_df, f)
        pkl.dump(id10, f)
        pkl.dump(indices, f)
        pkl.dump(fitted_y_train, f)
        pkl.dump(fitted_y_val, f)
        pkl.dump(fitted_y_test, f)

    gradients.to_csv("../output/"+output_folder+"/gradients_"+str(index)+".csv", index=False)
    gradients_masked.to_csv("../output/"+output_folder+"/gradients_masked_"+str(index)+".csv", index=False)


df1 = pd.DataFrame({'Station indices': indices})
df1.to_excel("../output/"+output_folder+"/station_indices.xlsx", sheet_name='sheet1', index=False,header=False)

'''
# testing the limits of the stationIDs
max_id = 0
min_id = 10000
for folder in listdir('../data/project2_energy_1812/raw/result'):
    if 'DS' not in folder and '._' not in folder: # address the DS_Store...
        for csv_file in listdir('../data/project2_energy_1812/raw/result/'+ folder):
            directory_path = '../data/project2_energy_1812/raw/result/'+folder+'/ID10.csv'
            pixelID = pd.read_csv(directory_path, header = None)
            pixelID = pixelID[((pixelID != 9999999.0) & (pixelID != 8888888.0))]
            if np.max(np.max(pixelID)) > max_id:
                max_id = np.max(np.max(pixelID))
                print(max_id)
            if np.min(np.min(pixelID)) < min_id:
                min_id = np.min(np.min(pixelID))
                print(min_id)
'''

